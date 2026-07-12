#!/usr/bin/env python3
"""Builds tools/reading-season.xlsx — the workbook readers score in.

Re-run any time: python3 tools/build-reading-template.py
Then upload to Google Sheets and import the exported applications CSV
into the Applications tab (instructions are on the workbook's first tab).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROSE = "B76E79"
ROSE_DEEP = "9C5560"
BLUSH = "F9ECEA"
INK = "33232A"
PAPER = "FBF7F5"
MAX_ROWS = 200  # formula coverage; plenty for one cycle

wb = Workbook()

thin = Side(style="thin", color="E9DCDA")
border = Border(bottom=thin)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=ROSE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color=ROSE_DEEP)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=11, color=INK)


# ---------- 1. How to use ----------
ws = wb.active
ws.title = "How to use"
ws.sheet_properties.tabColor = ROSE
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 110
title_block(ws, "", "")
ws["B2"] = "Rose Kelley Scholarship — Reading Season Workbook"
ws["B2"].font = Font(bold=True, size=18, color=ROSE_DEEP)
ws["B3"] = "Two readers, one published rubric, every applicant scored the same way."
ws["B3"].font = Font(italic=True, size=12, color=INK)

steps = [
    ("1.", "Export the applications: on the scholarship computer, run  node tools/export-applications.js"),
    ("2.", "Open the Applications tab, click cell A1 (the header row — the CSV brings its own), then File → Import → Upload the applications CSV → 'Replace data at selected cell'. The first applicant must land in row 2; if you ever see two header rows, delete the extra one."),
    ("3.", "Each reader scores every applicant on the Scoring tab — Reader 1 uses the rows marked Reader 1, Reader 2 likewise. Scores are whole numbers 1–5 (the tab enforces this). Do not discuss scores until both readers are done."),
    ("4.", "The Results tab fills itself in: it totals both readers per applicant and ranks them. Highest combined score receives the award."),
    ("5.", "Ties break by the 'What $1,000 changes' score, then a joint re-read (same as published on the site)."),
    ("6.", "A reader with any personal connection to an applicant skips those rows and a replacement reader scores them instead."),
    ("7.", "When the cycle ends: delete this workbook, the CSVs, and the database (delete-cycle-data.sh) — the privacy policy promises it."),
]
row = 5
for num, text in steps:
    ws[f"A{row}"] = num
    ws[f"A{row}"].font = Font(bold=True, color=ROSE_DEEP, size=12)
    ws[f"B{row}"] = text
    ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 32
    row += 1

ws[f"B{row+1}"] = "The rubric (1–5 points each, per reader)"
ws[f"B{row+1}"].font = Font(bold=True, size=12, color=ROSE_DEEP)
rubric = [
    "A path of your own — clarity and genuine interest in what they want to study, build, or become. Not ambition for its own sake.",
    "Steps already taken — anything done to move toward the path: classes, projects, jobs, self-teaching, caregiving, advocacy.",
    "What $1,000 changes — how concretely the award helps. Specific and honest beats dramatic.",
    "Writing style is NOT scored. Spelling, grammar, and polish earn zero points — substance does. Video/audio links score with the identical rubric.",
]
for i, t in enumerate(rubric):
    r = row + 2 + i
    ws[f"B{r}"] = ("• " if i < 3 else "⚠ ") + t
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30

# ---------- 2. Applications ----------
ws = wb.create_sheet("Applications")
ws.sheet_properties.tabColor = BLUSH
headers = ["ID", "Submitted", "Name", "Email", "School / program", "Where they are", "Essay"]
widths = [22, 22, 22, 28, 26, 22, 110]
for c, (h, w) in enumerate(zip(headers, widths), start=1):
    ws.cell(row=1, column=c, value=h)
    ws.column_dimensions[get_column_letter(c)].width = w
style_header(ws, len(headers))
ws.freeze_panes = "A2"
for r in range(2, MAX_ROWS + 2):
    ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = border

# ---------- 3. Scoring ----------
ws = wb.create_sheet("Scoring")
ws.sheet_properties.tabColor = ROSE_DEEP
headers = ["ID", "Applicant", "Reader",
           "A path of your own (1–5)", "Steps already taken (1–5)",
           "What $1,000 changes (1–5)", "Total", "Notes (private to readers)"]
widths = [22, 24, 12, 24, 24, 24, 10, 46]
for c, (h, w) in enumerate(zip(headers, widths), start=1):
    ws.cell(row=1, column=c, value=h)
    ws.column_dimensions[get_column_letter(c)].width = w
style_header(ws, len(headers))
ws.freeze_panes = "A2"

# Two rows per applicant, auto-pulled from Applications
for i in range(MAX_ROWS):
    app_row = 2 + i
    for j, reader in enumerate(("Reader 1", "Reader 2")):
        r = 2 + i * 2 + j
        blank = f'OR(Applications!A{app_row}="",Applications!A{app_row}="ID")'
        ws.cell(row=r, column=1, value=f'=IF({blank},"",Applications!A{app_row})')
        ws.cell(row=r, column=2, value=f'=IF({blank},"",Applications!C{app_row})')
        ws.cell(row=r, column=3, value=f'=IF({blank},"","{reader}")')
        ws.cell(row=r, column=7, value=f'=IF(COUNT(D{r}:F{r})=0,"",SUM(D{r}:F{r}))')
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical="top")

dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5,
                    allow_blank=True, showErrorMessage=True,
                    errorTitle="Scores are 1–5",
                    error="Each rubric category is scored 1 to 5, whole numbers only.")
ws.add_data_validation(dv)
dv.add(f"D2:F{1 + MAX_ROWS * 2}")

# ---------- 4. Results ----------
ws = wb.create_sheet("Results")
ws.sheet_properties.tabColor = INK
headers = ["ID", "Applicant", "Reader 1 total", "Reader 2 total", "Combined",
           "Tie-break: 'What $1,000 changes' combined", "Rank"]
widths = [22, 26, 14, 14, 12, 32, 8]
for c, (h, w) in enumerate(zip(headers, widths), start=1):
    ws.cell(row=1, column=c, value=h)
    ws.column_dimensions[get_column_letter(c)].width = w
style_header(ws, len(headers))
ws.freeze_panes = "A2"

for i in range(MAX_ROWS):
    r = 2 + i
    app_row = 2 + i
    r1 = 2 + i * 2       # Reader 1 row in Scoring
    r2 = 3 + i * 2       # Reader 2 row in Scoring
    blank = f'OR(Applications!A{app_row}="",Applications!A{app_row}="ID")'
    ws.cell(row=r, column=1, value=f'=IF({blank},"",Applications!A{app_row})')
    ws.cell(row=r, column=2, value=f'=IF({blank},"",Applications!C{app_row})')
    ws.cell(row=r, column=3, value=f'=IF({blank},"",Scoring!G{r1})')
    ws.cell(row=r, column=4, value=f'=IF({blank},"",Scoring!G{r2})')
    ws.cell(row=r, column=5, value=f'=IF({blank},"",IF(COUNT(C{r}:D{r})=0,"",SUM(C{r}:D{r})))')
    ws.cell(row=r, column=6, value=f'=IF({blank},"",SUM(Scoring!F{r1},Scoring!F{r2}))')
    ws.cell(row=r, column=7, value=f'=IF(OR({blank},E{r}=""),"",RANK(E{r},$E$2:$E${MAX_ROWS+1}))')
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = border

wb.save("tools/reading-season.xlsx")
print("wrote tools/reading-season.xlsx")
